import openpyxl
import os


def get_file():
    file_name_list = os.listdir()
    file_name_ = []
    for file_name in file_name_list:
        try:
            file, suffix = file_name.split('.')
            if suffix in ('xlsx', 'xlsm', 'xls') and '~$' not in file:
                print(file_name)
                file_name_.append(file_name)
        except:
            continue
    return file_name_


def handle_file(filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb.sheetnames[0]
    sheet2 = wb.sheetnames[1]
    ws1 = wb[sheet1]
    ws2 = wb[sheet2]

    data = {}

    for idx, row in enumerate(ws1.values):
        if idx == 0:
            continue
        if row[1] in data:
            data[row[1]][row[2]] = row[0]
        else:
            data[row[1]] = {row[2]: row[0]}
    cabin_list = list(data.keys())
    max_row = ws2.max_row
    max_column = ws2.max_column
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_column + 1):
            temp_row_idx = row_idx
            temp_col_idx = col_idx
            c = ws2.cell(row=row_idx, column=col_idx)
            cabin = c.value

            if cabin in cabin_list:
                flag = True
                while flag:
                    temp_row_idx += 1
                    server_cell = ws2.cell(row=temp_row_idx, column=temp_col_idx)
                    if server_cell.value not in cabin_list and temp_row_idx <= max_row:
                        u_pos = ws2.cell(row=temp_row_idx, column=(temp_col_idx - 2))
                        server_name = data.get(cabin, "").get(u_pos.value, "")
                        if server_name:
                            print(server_name, cabin, temp_row_idx, temp_col_idx)
                            ws2.cell(row=temp_row_idx, column=temp_col_idx, value=server_name)
                    else:
                        flag = False
    file_, _ = file_name.split('.')
    wb.save(f"{file_}_temp.xlsx")


if __name__ == '__main__':

    file_name_list = get_file()
    for file_name in file_name_list:
        try:
            handle_file(file_name)
        except:
            continue
